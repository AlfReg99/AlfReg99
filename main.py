import codecs
import bs4
from Data import getFamilyNameFromFamilyId, getGroupNameFromGroupId
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

def ReadFromHTML(pathfile):
    file = codecs.open(pathfile, "r", "utf-8")
    return file.read()

pathfile=r"C:\Users\alfon\Desktop\report.html"

htmlfile=ReadFromHTML(pathfile)

body = bs4.BeautifulSoup(htmlfile, 'lxml').find('body')
firstdiv=body.find('div')
detailedContainer=firstdiv.find('div', {'id': 'detailsContainer'})
assessmentDetailedArea=detailedContainer.find('div', {'id': 'assessmentDetailsArea'})
benchmarksList=assessmentDetailedArea.find_all('div', {'class': 'Rule'})

workbook = Workbook()

worksheet = workbook.active
worksheet.title = 'Report'
columns = ['Famiglia', 'Gruppo', 'CIS Benchmark', 'Descrizione','CIS Controls', 'CIS Subcontrols', 'Implementation Group', 'Risultato Atteso', 'Risultato Ottenuto' , 'Esito']
row_num = 1

for col_num, column_title in enumerate(columns, 1):
    cell = worksheet.cell(row=row_num, column=col_num)
    cell.value = column_title
    cell.font = Font(name="Times New Roman",size=12,bold=True,color='FF0000')
    cell.border = Border(left=Side(border_style="thin",color='FF000000'),
                         right=Side(border_style="thin",color='FF000000'),
                         top=Side(border_style="thin",color='FF000000'),
                         bottom=Side(border_style="thin",color='FF000000'),)

for benchmarks in benchmarksList:
    row_num += 1

    try:
        result = benchmarks.find('span').contents[0]
    except:
        result = ''
    try:
        benchmark = benchmarks.find('h3').contents[0]
    except:
        benchmark = '-'
    try:
        description = benchmarks.find('p').contents[0]
    except:
        description = ""

    cisControls = benchmarks.find_all('b')
    sectionsCisControls = benchmarks.find_all('div', {'class': 'cceevidence rule-xml'})

    try:
        familyId = benchmark.split(".")[0]
        family=getFamilyNameFromFamilyId(familyId)
    except:
        family="N/A"
    try:
        benchmarkSplitted = benchmark.split(".")
        groupdId = benchmarkSplitted[0] + "." + benchmarkSplitted[1]
        group=getGroupNameFromGroupId(groupdId)
    except:
        group="N/A"

    try:
        cisControlString = ""
        for cisControl in cisControls:
            cisControlString += str(cisControl.contents[0]) + " "
    except:
        cisControlString = "N/A"


    igGroup = "N/A"
    subcontrols=""
    try:
        for section in sectionsCisControls:
            subcontrols = section.find_all('td')[3].contents[0]
            print('Subcontrols: ' + subcontrols)

            widthList = section.find_all('td', {'width': '80%'})
            if (len(widthList) > 4):
                igGroup = widthList[4].contents[0]
                print("IGs: " + igGroup)
    except:
        igGroup = "N/A"
        subcontrols= "N/A"

    try:
        expected = benchmarks.find('div', {'class': 'check'}).find('div', {'class': 'xml'}).find('table').find(
            'tr').find('td', {'width': '95%'}) \
            .find('div').find('div').find('div').find('table', {'class': 'evidence-sep'}).find('tbody').find('tr')
        expectedResult = expected.find_all('td')[1].contents[0]
    except:
        expectedResult = "N/A"

    try:
        obtained = benchmarks.find('div', {'class': 'check'}).find('div', {'class': 'xml'}).find('table').find(
            'tr').find('td', {'width': '95%'}) \
            .find('div').find('div').find('div').find('table', {'class': 'evidence'}).find('tbody').find('tr', {
            'class': 'evaluated'})
        obtainedResult = obtained.find_all('td')[3].contents[0]
    except:
        obtainedResult = "N/A"


    row = [
        family,
        group,
        benchmark,
        description,
        cisControlString,
        subcontrols,
        igGroup,
        expectedResult,
        obtainedResult,
        result
    ]

    # Assign the data for each cell of the row
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
        cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                             right=Side(border_style="thin", color='FF000000'),
                             top=Side(border_style="thin", color='FF000000'),
                             bottom=Side(border_style="thin", color='FF000000'), )

    count_attr = 0
    old_row = row_num
    print('\n\n')

workbook.save(filename=r"C:\Users\alfon\Desktop\Progetti\script_tabella\venv1\results\results.xlsx")
workbook.close()